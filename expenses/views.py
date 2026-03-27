"""
Xarajatlar - Dashboard, CRUD, sozlamalar.
"""
import logging
import os
import re
import secrets
from datetime import datetime
from tempfile import NamedTemporaryFile

import requests
from django.conf import settings
from django.core.cache import cache

logger = logging.getLogger(__name__)

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_http_methods

from django.db.models import F, ExpressionWrapper, DecimalField

from accounts.models import FinanceProfile
from categories.services import create_default_categories
from core.permissions import get_user_object_or_404
from core.rate_limit import rate_limit_action
from .models import Expense, SavingGoal, RecurringExpense, Debt
from .currency import get_currency_rates_to_uzs
from .forms import ExpenseForm, SavingGoalForm, RecurringExpenseForm, DebtForm
from .services import advance_next_payment, get_dashboard_context, get_monthly_totals, invalidate_monthly_totals_cache
from analytics.services import get_insights_for_user, get_user_achievements, get_month_end_forecast
from notifications.services import (
    maybe_send_limit_warning_after_expense,
    maybe_send_expense_confirmation_after_expense,
    send_telegram_message,
    send_telegram_document,
)


MONTH_NAMES = [
    "", "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
    "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr",
]
OCR_SUGGESTION_CACHE_KEY = "ocr_suggestion:{user_id}:{donation_id}"


def _extract_amount_from_ocr_text(text: str) -> int:
    """OCR matnidan eng ehtimoliy donat summani topadi."""
    raw = (text or "").strip()
    if not raw:
        return 0
    candidates = re.findall(r"[\d][\d\s,\.]{2,}", raw)
    best = 0
    for item in candidates:
        digits = re.sub(r"[^\d]", "", item)
        if not digits:
            continue
        try:
            val = int(digits)
        except ValueError:
            continue
        if 1000 <= val <= 500_000_000:
            best = max(best, val)
    return best


def _ocr_extract_amount_for_donation(donation) -> tuple[int, str]:
    """
    Telegram screenshot + OCR.Space API orqali summani aniqlashga urinadi.
    OCRSPACE_API_KEY bo'lmasa xatolik qaytaradi.
    """
    token = getattr(settings, "TELEGRAM_BOT_TOKEN", "")
    api_key = (getattr(settings, "OCRSPACE_API_KEY", "") or "").strip()
    if not token:
        return 0, "Bot token sozlanmagan."
    if not api_key:
        return 0, "OCRSPACE_API_KEY sozlanmagan."
    if not donation or not donation.screenshot_file_id:
        return 0, "Screenshot topilmadi."

    try:
        get_file_url = f"https://api.telegram.org/bot{token}/getFile"
        meta_resp = requests.get(get_file_url, params={"file_id": donation.screenshot_file_id}, timeout=10)
        meta_resp.raise_for_status()
        file_path = ((meta_resp.json() or {}).get("result") or {}).get("file_path")
        if not file_path:
            return 0, "Telegram fayl yo'li topilmadi."
        file_url = f"https://api.telegram.org/file/bot{token}/{file_path}"
        img_resp = requests.get(file_url, timeout=20)
        img_resp.raise_for_status()

        ocr_resp = requests.post(
            "https://api.ocr.space/parse/image",
            data={
                "apikey": api_key,
                "language": "eng",
                "OCREngine": "2",
                "isOverlayRequired": "false",
                "scale": "true",
            },
            files={"filename": ("donation.jpg", img_resp.content)},
            timeout=30,
        )
        ocr_resp.raise_for_status()
        payload = ocr_resp.json() or {}
        parsed_list = payload.get("ParsedResults") or []
        text = "\n".join((item or {}).get("ParsedText", "") for item in parsed_list).strip()
        amount = _extract_amount_from_ocr_text(text)
        if amount <= 0:
            return 0, "OCR orqali summa aniqlanmadi."
        return amount, ""
    except Exception as e:
        logger.warning("donation OCR failed donation_id=%s: %s", getattr(donation, "pk", None), e)
        return 0, "OCR xizmati bilan ishlashda xatolik."


def _generate_household_invite_code() -> str:
    return secrets.token_hex(4).upper()


def _safe_next_url(request, fallback="expenses:dashboard"):
    """Faqat shu host ichidagi next URL ni qabul qiladi."""
    next_url = request.POST.get("next") or request.GET.get("next")
    if next_url and url_has_allowed_host_and_scheme(
        url=next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return next_url
    return fallback


@login_required
def dashboard(request):
    """Bu Oy - asosiy dashboard. Ma'lumot get_dashboard_context orqali yig'iladi."""
    profile, _ = FinanceProfile.objects.get_or_create(user=request.user)
    if not profile.onboarding_completed:
        return redirect("expenses:onboarding")

    today = timezone.now().date()
    date_str = request.GET.get("date") or ""
    selected_date = today
    if date_str:
        try:
            selected_date = datetime.fromisoformat(date_str).date()
        except ValueError:
            selected_date = today

    try:
        context = get_dashboard_context(request.user, selected_date)
    except Exception as e:
        logger.exception("dashboard get_dashboard_context user_id=%s: %s", request.user.pk, e)
        messages.error(
            request,
            "Dashboard ma'lumotlarini yuklashda xatolik. Sahifani yangilab ko'ring.",
        )
        net_debt = 0
        context = {
            "totals": {"total_spent": 0, "budget": 0, "remaining": 0, "month_start": today, "month_end": today, "year": today.year, "month": today.month},
            "month_display": f"{MONTH_NAMES[today.month]} {today.year}",
            "selected_date_display": f"{today.day} {MONTH_NAMES[today.month].lower()} {today.year}",
            "selected_date_iso": today.isoformat(),
            "breakdown": [],
            "recent": [],
            "avg_daily": 0,
            "top_categories": [],
            "daily_summary": None,
            "upcoming_recurring": [],
            "net_debt": net_debt,
            "net_debt_abs": abs(net_debt),
            "taken_debt_total": 0,
            "given_debt_total": 0,
        }
    context["finance_profile"] = profile
    context["needs_budget"] = not bool(getattr(request.user, "monthly_budget", 0))
    try:
        context["insights"] = get_insights_for_user(
            request.user,
            year=selected_date.year,
            month=selected_date.month,
            limit=5,
        )
    except Exception:
        context["insights"] = []
    context["can_see_full_insights"] = getattr(request.user, "is_supporter", False)
    try:
        context["achievements"] = get_user_achievements(request.user, limit=3)
    except Exception:
        context["achievements"] = []
    context["household_summary"] = None
    if request.user.household_id and request.user.household_share_data and request.user.is_supporter:
        month_start = context["totals"]["month_start"]
        month_end = context["totals"]["month_end"]
        household_qs = Expense.objects.filter(
            user__household_id=request.user.household_id,
            user__household_share_data=True,
            date__gte=month_start,
            date__lte=month_end,
        )
        from django.db.models import Sum

        household_total = household_qs.aggregate(s=Sum("amount"))["s"] or 0
        household_members = (
            request.user.household.members.filter(household_share_data=True).count()
        )
        context["household_summary"] = {
            "name": request.user.household.name,
            "total_spent": household_total,
            "members": household_members,
        }
    try:
        context["month_forecast"] = get_month_end_forecast(
            request.user,
            year=selected_date.year,
            month=selected_date.month,
        )
    except Exception:
        context["month_forecast"] = None
    try:
        rates = get_currency_rates_to_uzs()
        usd_rate = rates.get("USD")
        context["usd_rate_uzs"] = int(usd_rate) if usd_rate else None
        from .models import ExchangeRate

        latest_rate_date = ExchangeRate.objects.order_by("-date").values_list("date", flat=True).first()
        context["usd_rate_date"] = latest_rate_date
        context["usd_rate_source"] = "db" if latest_rate_date else "env"
    except Exception:
        context["usd_rate_uzs"] = None
        context["usd_rate_date"] = None
        context["usd_rate_source"] = "env"
    return render(request, "expenses/dashboard.html", context)


@login_required
@require_http_methods(["GET", "POST"])
def onboarding_view(request):
    """Birinchi ishga tushirish uchun oddiy moliyaviy onboarding."""
    user = request.user
    profile, _ = FinanceProfile.objects.get_or_create(user=user)

    if profile.onboarding_completed and request.method == "GET":
        return redirect("expenses:dashboard")

    if request.method == "POST":
        if request.POST.get("skip_budget") == "1":
            profile.onboarding_completed = True
            profile.save(update_fields=["onboarding_completed"])
            create_default_categories(user)
            messages.success(request, "Sozlamalarni keyinroq Sozlamalar bo‘limida to‘ldirishingiz mumkin.")
            return redirect("expenses:dashboard")

        monthly_budget = request.POST.get("monthly_budget")
        primary_goal = request.POST.get("primary_goal") or profile.primary_goal

        if monthly_budget is not None and monthly_budget.strip() != "":
            try:
                normalized_budget = "".join(ch for ch in monthly_budget if ch.isdigit())
                user.monthly_budget = int(normalized_budget) if normalized_budget else 0
            except (ValueError, TypeError):
                pass
            user.save(update_fields=["monthly_budget"])
        else:
            user.monthly_budget = 0
            user.save(update_fields=["monthly_budget"])

        profile.primary_goal = primary_goal
        profile.onboarding_completed = True
        profile.save(update_fields=["primary_goal", "onboarding_completed"])
        create_default_categories(user)

        messages.success(request, "Asosiy moliyaviy sozlamalar saqlandi.")
        return redirect("expenses:dashboard")

    return render(
        request,
        "expenses/onboarding.html",
        {
            "user": user,
            "profile": profile,
            "goal_choices": FinanceProfile.PrimaryGoal.choices,
        },
    )


GOALS_PAGE_SIZE = 25


@login_required
def saving_goal_list(request):
    """Foydalanuvchining jamg'arma maqsadlari (pagination)."""
    goals = (
        SavingGoal.objects.filter(user=request.user)
        .annotate(
            remaining=ExpressionWrapper(
                F("target_amount") - F("current_amount"),
                output_field=DecimalField(),
            )
        )
        .order_by("-is_active", "remaining", "-created_at")
    )
    paginator = Paginator(goals, GOALS_PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    active_goals = [g for g in page_obj if g.is_active]
    archived_goals = [g for g in page_obj if not g.is_active]
    return render(
        request,
        "expenses/saving_goal_list.html",
        {
            "page_obj": page_obj,
            "active_goals": active_goals,
            "archived_goals": archived_goals,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
@rate_limit_action("saving_goal_create", max_requests=30)
def saving_goal_create(request):
    """Yangi jamg'arma maqsadi yaratish."""
    form = SavingGoalForm(request.POST or None, user=request.user)
    if form.is_valid():
        form.save()
        messages.success(request, "Yangi jamg'arma maqsadi saqlandi.")
        return redirect("expenses:saving_goal_list")
    return render(
        request,
        "expenses/saving_goal_form.html",
        {
            "form": form,
            "title": "Yangi maqsad",
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
@rate_limit_action("saving_goal_edit", max_requests=30)
def saving_goal_edit(request, pk):
    """Mavjud jamg'arma maqsadini tahrirlash."""
    goal = get_user_object_or_404(SavingGoal, request.user, pk)
    form = SavingGoalForm(request.POST or None, instance=goal, user=request.user)
    if form.is_valid():
        form.save()
        messages.success(request, "Maqsad yangilandi.")
        return redirect("expenses:saving_goal_list")
    return render(
        request,
        "expenses/saving_goal_form.html",
        {
            "form": form,
            "goal": goal,
            "title": "Maqsadni tahrirlash",
        },
    )


RECURRING_PAGE_SIZE = 25


@login_required
def recurring_list(request):
    """Qayta takrorlanuvchi chiqimlar ro'yxati (pagination)."""
    today = timezone.now().date()
    qs = (
        RecurringExpense.objects.filter(user=request.user)
        .order_by("-is_active", "next_payment_date", "-updated_at")
    )
    paginator = Paginator(qs, RECURRING_PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    upcoming = [r for r in page_obj if r.is_active and r.next_payment_date >= today]
    archived = [r for r in page_obj if not r.is_active]
    return render(
        request,
        "expenses/recurring_list.html",
        {
            "page_obj": page_obj,
            "upcoming": upcoming,
            "archived": archived,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
@rate_limit_action("recurring_create", max_requests=20)
def recurring_create(request):
    """Yangi qayta takrorlanuvchi chiqim yaratish."""
    form = RecurringExpenseForm(request.POST or None, user=request.user)
    if form.is_valid():
        form.save()
        messages.success(request, "Qayta takrorlanuvchi chiqim saqlandi.")
        return redirect("expenses:recurring_list")
    return render(
        request,
        "expenses/recurring_form.html",
        {
            "form": form,
            "title": "Yangi qayta chiqim",
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
@rate_limit_action("recurring_edit", max_requests=30)
def recurring_edit(request, pk):
    """Mavjud qayta takrorlanuvchi chiqimni tahrirlash."""
    obj = get_user_object_or_404(RecurringExpense, request.user, pk)
    form = RecurringExpenseForm(request.POST or None, instance=obj, user=request.user)
    if form.is_valid():
        form.save()
        messages.success(request, "Qayta takrorlanuvchi chiqim yangilandi.")
        return redirect("expenses:recurring_list")
    return render(
        request,
        "expenses/recurring_form.html",
        {
            "form": form,
            "recurring": obj,
            "title": "Qayta chiqimni tahrirlash",
        },
    )


@login_required
@require_http_methods(["POST"])
@rate_limit_action("recurring_delete", max_requests=30)
def recurring_delete(request, pk):
    obj = get_user_object_or_404(RecurringExpense, request.user, pk)
    obj.delete()
    messages.success(request, "Qayta takrorlanuvchi chiqim o'chirildi.")
    return redirect("expenses:recurring_list")


@login_required
@require_http_methods(["POST"])
@rate_limit_action("recurring_mark_paid", max_requests=15)
def recurring_mark_paid(request, pk):
    """To'lov qilindi: keyingi sana yangilanadi, ixtiyoriy Expense yaratiladi."""
    obj = get_user_object_or_404(RecurringExpense, request.user, pk)
    create_expense = request.POST.get("create_expense") == "1"
    try:
        advance_next_payment(obj, create_expense=create_expense)
        messages.success(
            request,
            "To'lov qabul qilindi. Keyingi to'lov sanasi yangilandi."
            + (" Xarajat yozuvi qo'shildi." if create_expense else ""),
        )
    except Exception as e:
        logger.exception("recurring_mark_paid error pk=%s: %s", pk, e)
        messages.error(request, "Xatolik yuz berdi. Keyinroq qayta urinib ko'ring.")
    return redirect("expenses:recurring_list")


DEBTS_PAGE_SIZE = 25


@login_required
def debt_list(request):
    """Qarz va qarzdorliklar ro'yxati (pagination)."""
    from decimal import Decimal
    from django.db.models import Sum

    debts_qs = (
        Debt.objects.filter(user=request.user)
        .order_by("is_closed", "due_date", "-created_at")
    )
    paginator = Paginator(debts_qs, DEBTS_PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    open_debts = [d for d in page_obj if not d.is_closed]
    closed_debts = [d for d in page_obj if d.is_closed]
    debt_base = Debt.objects.filter(user=request.user, is_closed=False)
    taken_total = debt_base.filter(kind=Debt.Kind.TAKEN).aggregate(s=Sum("amount"))["s"] or Decimal("0")
    given_total = debt_base.filter(kind=Debt.Kind.GIVEN).aggregate(s=Sum("amount"))["s"] or Decimal("0")
    net_debt = taken_total - given_total
    net_debt_abs = abs(net_debt)
    return render(
        request,
        "expenses/debt_list.html",
        {
            "page_obj": page_obj,
            "open_debts": open_debts,
            "closed_debts": closed_debts,
            "taken_debt_total": taken_total,
            "given_debt_total": given_total,
            "net_debt": net_debt,
            "net_debt_abs": net_debt_abs,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
@rate_limit_action("debt_create", max_requests=20)
def debt_create(request):
    """Yangi qarz yoki qarzdorlik yozuvi yaratish."""
    form = DebtForm(request.POST or None, user=request.user)
    if form.is_valid():
        form.save()
        messages.success(request, "Qarz yozuvi saqlandi.")
        return redirect("expenses:debt_list")
    return render(
        request,
        "expenses/debt_form.html",
        {
            "form": form,
            "title": "Yangi qarz yozuvi",
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
@rate_limit_action("debt_edit", max_requests=30)
def debt_edit(request, pk):
    """Mavjud qarz yozuvini tahrirlash yoki yopish."""
    obj = get_user_object_or_404(Debt, request.user, pk)
    form = DebtForm(request.POST or None, instance=obj, user=request.user)
    if form.is_valid():
        form.save()
        messages.success(request, "Qarz yozuvi yangilandi.")
        return redirect("expenses:debt_list")
    return render(
        request,
        "expenses/debt_form.html",
        {
            "form": form,
            "debt": obj,
            "title": "Qarz yozuvini tahrirlash",
        },
    )


@login_required
@require_http_methods(["POST"])
@rate_limit_action("debt_delete", max_requests=30)
def debt_delete(request, pk):
    obj = get_user_object_or_404(Debt, request.user, pk)
    obj.delete()
    messages.success(request, "Qarz yozuvi o'chirildi.")
    return redirect("expenses:debt_list")


@login_required
@require_http_methods(["GET", "POST"])
@rate_limit_action("expense_add", max_requests=25)
def expense_add(request):
    template_id_raw = (request.GET.get("tpl") or "").strip()
    initial_data = None
    if request.method == "GET" and template_id_raw.isdigit():
        tpl_obj = (
            Expense.objects.filter(user=request.user, pk=int(template_id_raw))
            .select_related("category")
            .first()
        )
        if tpl_obj:
            initial_data = {
                "category": tpl_obj.category_id,
                "amount": int(tpl_obj.original_amount or tpl_obj.amount or 0),
                "currency": tpl_obj.currency or "UZS",
                "note": tpl_obj.note or "",
                "date": timezone.now().date(),
            }
    form = ExpenseForm(request.POST or None, user=request.user, initial=initial_data)
    if form.is_valid():
        try:
            expense = form.save()
            invalidate_monthly_totals_cache(request.user, year=expense.date.year, month=expense.date.month)
            maybe_send_limit_warning_after_expense(request.user)
            maybe_send_expense_confirmation_after_expense(request.user, expense)
            add_another = request.POST.get("add_another") == "1"
            messages.success(
                request,
                "Xarajat qo'shildi. Yana bittasini tezda qo'shishingiz mumkin."
                if add_another
                else "Xarajat qo'shildi.",
            )
            logger.info("expense_add user_id=%s amount=%s date=%s", request.user.pk, expense.amount, expense.date)
            if add_another:
                return redirect("expenses:add")
            return redirect(_safe_next_url(request))
        except Exception as e:
            logger.exception("expense_add save user_id=%s: %s", request.user.pk, e)
            messages.error(request, "Xarajatni saqlashda xatolik. Qaytadan urinib ko'ring.")
    recent_templates = (
        Expense.objects.filter(user=request.user)
        .select_related("category")
        .order_by("-created_at")[:5]
    )
    return render(
        request,
        "expenses/expense_form.html",
        {"form": form, "title": "Xarajat qo'shish", "recent_templates": recent_templates},
    )


@login_required
@require_http_methods(["GET", "POST"])
@rate_limit_action("expense_edit", max_requests=30)
def expense_edit(request, pk):
    expense = get_user_object_or_404(Expense, request.user, pk)
    form = ExpenseForm(request.POST or None, instance=expense, user=request.user)
    if form.is_valid():
        expense = form.save()
        invalidate_monthly_totals_cache(request.user, year=expense.date.year, month=expense.date.month)
        maybe_send_limit_warning_after_expense(request.user)
        maybe_send_expense_confirmation_after_expense(request.user, expense)
        messages.success(request, "Xarajat yangilandi.")
        return redirect(_safe_next_url(request))
    return render(request, "expenses/expense_form.html", {"form": form, "expense": expense, "title": "Xarajatni tahrirlash"})


@login_required
@require_http_methods(["POST"])
@rate_limit_action("expense_delete", max_requests=30)
def expense_delete(request, pk):
    expense = get_user_object_or_404(Expense, request.user, pk)
    year, month = expense.date.year, expense.date.month
    expense.delete()
    invalidate_monthly_totals_cache(request.user, year=year, month=month)
    messages.success(request, "Xarajat o'chirildi.")
    return redirect(_safe_next_url(request))


@login_required
def expense_list(request):
    """Barcha xarajatlar (pagination, oy filtri, tez oraliq filtri va qidiruv)."""
    today = timezone.now().date()
    scope = (request.GET.get("scope") or "my").strip().lower()
    if scope not in {"my", "household"}:
        scope = "my"
    if (
        scope == "household"
        and request.user.household_id
        and request.user.is_supporter
        and request.user.household_share_data
    ):
        qs = Expense.objects.filter(
            user__household_id=request.user.household_id,
            user__household_share_data=True,
        ).select_related("category", "user").order_by("-date", "-created_at")
    else:
        scope = "my"
        qs = Expense.objects.filter(user=request.user).select_related("category").order_by("-date", "-created_at")

    # Matnli qidiruv: izoh va turkum nomi bo'yicha
    q = (request.GET.get("q") or "").strip()
    if q:
        from django.db.models import Q

        qs = qs.filter(
            Q(note__icontains=q) | Q(category__name__icontains=q)
        )

    # Tez vaqt oralig'i filtri (month/year filtridan oldin qo'llanadi)
    range_key = (request.GET.get("range") or "").strip().lower()
    if range_key == "today":
        qs = qs.filter(date=today)
    elif range_key == "7d":
        start_7d = today - timezone.timedelta(days=6)
        qs = qs.filter(date__gte=start_7d, date__lte=today)
    elif range_key == "month":
        qs = qs.filter(date__year=today.year, date__month=today.month)

    year_str = request.GET.get("year")
    month_str = request.GET.get("month")
    if year_str and month_str and range_key not in {"today", "7d", "month"}:
        try:
            year = int(year_str)
            month = int(month_str)
            if 1 <= month <= 12 and 2000 <= year <= 2100:
                from calendar import monthrange
                _, last_day = monthrange(year, month)
                from datetime import date
                start = date(year, month, 1)
                end = date(year, month, last_day)
                qs = qs.filter(date__gte=start, date__lte=end)
            else:
                year = None
                month = None
        except (ValueError, TypeError):
            year = None
            month = None
    else:
        year = None
        month = None

    paginator = Paginator(qs, 20)
    page = request.GET.get("page", 1)
    page_obj = paginator.get_page(page)

    # Oxirgi 12 oy + "Barchasi" (filtr uchun)
    period_choices = []
    for i in range(12):
        m = today.month - i
        y = today.year
        while m <= 0:
            m += 12
            y -= 1
        period_choices.append((y, m, f"{MONTH_NAMES[m]} {y}"))

    # Pagination linklari uchun GET parametrlar (page dan tashqari)
    get_copy = request.GET.copy()
    if "page" in get_copy:
        get_copy.pop("page")
    base_query = get_copy.urlencode()

    return render(
        request,
        "expenses/expense_list.html",
        {
            "page_obj": page_obj,
            "period_choices": period_choices,
            "selected_year": year,
            "selected_month": month,
            "selected_range": range_key,
            "query_text": q,
            "base_query": base_query,
            "scope": scope,
        },
    )


EXPORT_MAX_MONTHS = 24  # Eksportda faqat oxirgi N oy


@login_required
@rate_limit_action("export_csv", max_requests=10, window=600)
def export_view(request):
    """CSV eksport (oxirgi 24 oy)."""
    import csv
    from django.http import HttpResponse

    today = timezone.now().date()
    cutoff = today - timezone.timedelta(days=EXPORT_MAX_MONTHS * 31)
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="chiqimlar.csv"'
    response.write("\ufeff")  # BOM for Excel UTF-8
    writer = csv.writer(response)
    writer.writerow(["Sana", "Turkum", "Summa", "Izoh"])
    qs = (
        Expense.objects.filter(user=request.user, date__gte=cutoff)
        .select_related("category")
        .order_by("-date", "-created_at")
    )
    for e in qs:
        writer.writerow([e.date, e.category.name if e.category else "", e.amount, e.note or ""])
    return response


@login_required
@require_http_methods(["POST"])
@rate_limit_action("export_excel_telegram", max_requests=10, window=600)
def export_excel_to_telegram(request):
    """
    Excel faylni yaratib, Telegram bot orqali yuboradi.
    Varaqlar: Chiqimlar (joriy oy), Byudjet, Qarzlar, Maqsadlar.
    """
    user = request.user
    if not user.telegram_id:
        messages.error(request, "Telegram hisob topilmadi. Botda /start yuborib qayta urinib ko'ring.")
        return redirect("expenses:settings")

    tmp_file_path = ""
    try:
        from openpyxl import Workbook

        today = timezone.now().date()
        month_start = today.replace(day=1)
        import calendar
        last_day = calendar.monthrange(today.year, today.month)[1]
        month_end = today.replace(day=last_day)

        wb = Workbook(write_only=True)

        # 1) Chiqimlar — joriy oy
        ws_exp = wb.create_sheet(title="Chiqimlar")
        ws_exp.append(["Sana", "Turkum", "Summa", "Izoh"])
        qs_exp = (
            Expense.objects.filter(user=user, date__gte=month_start, date__lte=month_end)
            .select_related("category")
            .order_by("-date", "-created_at")
        )
        for e in qs_exp.iterator(chunk_size=500):
            ws_exp.append([str(e.date), e.category.name if e.category else "", int(e.amount), e.note or ""])

        # 2) Byudjet
        totals = get_monthly_totals(user, year=today.year, month=today.month)
        ws_budget = wb.create_sheet(title="Byudjet")
        ws_budget.append(["Ko'rsatkich", "Qiymat (so'm)"])
        budget_val = totals.get("budget") or 0
        spent_val = totals.get("total_spent") or 0
        remaining_val = budget_val - spent_val
        ws_budget.append(["Oylik limit", int(budget_val)])
        ws_budget.append(["Joriy oy sarflangan", int(spent_val)])
        ws_budget.append(["Qolgan", int(remaining_val)])

        # 3) Qarzlar
        ws_debt = wb.create_sheet(title="Qarzlar")
        ws_debt.append(["Turi", "Kimga / Kimdan", "Summa", "Sana", "Qaytarish muddati", "Izoh", "Yopilgan"])
        for d in Debt.objects.filter(user=user).order_by("-is_closed", "due_date"):
            ws_debt.append([
                d.get_kind_display(),
                d.counterparty or "",
                int(d.amount),
                str(d.date),
                str(d.due_date) if d.due_date else "",
                d.note or "",
                "Ha" if d.is_closed else "Yo'q",
            ])

        # 4) Maqsadlar
        ws_goals = wb.create_sheet(title="Maqsadlar")
        ws_goals.append(["Nomi", "Maqsad summa", "Joriy summa", "Foiz %", "Boshlanish", "Tugash", "Faol"])
        for g in SavingGoal.objects.filter(user=user).order_by("-is_active", "-created_at"):
            ws_goals.append([
                g.name or "",
                int(g.target_amount),
                int(g.current_amount),
                g.progress_percent,
                str(g.start_date),
                str(g.target_date) if g.target_date else "",
                "Ha" if g.is_active else "Yo'q",
            ])

        with NamedTemporaryFile(prefix=f"chiqimlar_{user.pk}_", suffix=".xlsx", delete=False) as tmp:
            tmp_file_path = tmp.name
        wb.save(tmp_file_path)

        caption = f"📥 {today} holatiga: chiqimlar, byudjet, qarzlar, maqsadlar."
        ok = send_telegram_document(user.telegram_id, tmp_file_path, caption=caption)
        if ok:
            messages.success(request, "Excel fayl Telegram bot orqali yuborildi.")
        else:
            messages.error(request, "Excel yuborilmadi. Botga /start yuborib qayta urinib ko'ring.")
    except Exception as e:
        logger.exception("export_excel_to_telegram user_id=%s: %s", getattr(user, "pk", None), e)
        if settings.DEBUG:
            raise
        messages.error(request, "Excel eksportda xatolik yuz berdi. Keyinroq qayta urinib ko'ring.")
    finally:
        if tmp_file_path and os.path.exists(tmp_file_path):
            try:
                os.remove(tmp_file_path)
            except OSError:
                pass

    return redirect("expenses:settings")


@login_required
@require_http_methods(["GET", "POST"])
@rate_limit_action("settings_save", max_requests=30)
def settings_view(request):
    """Sozlamalar - oylik limit, kategoriyalar, eksport, donat/supporter."""
    from accounts.models import Donation, DonationMethod

    user = request.user
    profile, _ = FinanceProfile.objects.get_or_create(user=user)
    if request.method == "POST":
        action = (request.POST.get("action") or "save_settings").strip()
        if action == "household_create":
            from accounts.models import Household
            if not user.is_supporter:
                messages.error(request, "Family/Household funksiyasi faqat Donater foydalanuvchilar uchun.")
                return redirect("expenses:settings")

            name = (request.POST.get("household_name") or "").strip()
            if not name:
                messages.error(request, "Household nomini kiriting.")
                return redirect("expenses:settings")
            if user.household_id:
                messages.info(request, "Siz allaqachon householdga ulangan ekansiz.")
                return redirect("expenses:settings")
            invite_code = _generate_household_invite_code()
            while Household.objects.filter(invite_code=invite_code).exists():
                invite_code = _generate_household_invite_code()
            h = Household.objects.create(name=name, owner=user, invite_code=invite_code)
            user.household = h
            user.household_share_data = True
            user.save(update_fields=["household", "household_share_data"])
            messages.success(request, f"Household yaratildi. Taklif kodi: {invite_code}")
            return redirect("expenses:settings")
        if action == "household_join":
            from accounts.models import Household
            if not user.is_supporter:
                messages.error(request, "Family/Household funksiyasi faqat Donater foydalanuvchilar uchun.")
                return redirect("expenses:settings")

            code = (request.POST.get("invite_code") or "").strip().upper()
            if not code:
                messages.error(request, "Taklif kodini kiriting.")
                return redirect("expenses:settings")
            h = Household.objects.filter(invite_code=code).first()
            if not h:
                messages.error(request, "Taklif kodi topilmadi.")
                return redirect("expenses:settings")
            user.household = h
            user.household_share_data = True
            user.save(update_fields=["household", "household_share_data"])
            messages.success(request, f"Householdga qo'shildingiz: {h.name}")
            return redirect("expenses:settings")
        if action == "household_leave":
            if not user.is_supporter:
                messages.error(request, "Family/Household funksiyasi faqat Donater foydalanuvchilar uchun.")
                return redirect("expenses:settings")
            if user.household_id:
                user.household = None
                user.household_share_data = False
                user.save(update_fields=["household", "household_share_data"])
                messages.success(request, "Householddan chiqdingiz.")
            return redirect("expenses:settings")

        monthly_budget_raw = request.POST.get("monthly_budget")
        budget_ok = True
        if monthly_budget_raw is not None and monthly_budget_raw.strip() != "":
            try:
                normalized_budget = "".join(ch for ch in monthly_budget_raw if ch.isdigit())
                val = int(normalized_budget) if normalized_budget else 0
                if val < 0:
                    messages.error(request, "Oylik byudjet manfiy bo‘lmasligi kerak.")
                    budget_ok = False
                else:
                    user.monthly_budget = val
            except (ValueError, TypeError):
                messages.error(request, "Oylik byudjet uchun to‘g‘ri son kiriting.")
                budget_ok = False
        else:
            user.monthly_budget = 0
        user.telegram_notifications = request.POST.get("telegram_notifications") == "on"
        user.daily_reminder = request.POST.get("daily_reminder") == "on"
        user.weekly_summary = request.POST.get("weekly_summary") == "on"
        user.limit_warning = request.POST.get("limit_warning") == "on"
        user.household_share_data = request.POST.get("household_share_data") == "on"
        preferred_currency = (request.POST.get("preferred_currency") or "UZS").strip().upper()
        if preferred_currency not in {"UZS", "USD", "EUR", "RUB"}:
            preferred_currency = "UZS"
        profile.preferred_currency = preferred_currency
        profile.save(update_fields=["preferred_currency"])
        if budget_ok:
            user.save(
                update_fields=[
                    "monthly_budget",
                    "telegram_notifications",
                    "daily_reminder",
                    "weekly_summary",
                    "limit_warning",
                    "household_share_data",
                ]
            )
            messages.success(request, "Sozlamalar saqlandi.")
        return redirect("expenses:settings")
    donation_methods = DonationMethod.objects.filter(is_active=True).order_by("sort_order", "id")
    pending_donation_exists = Donation.objects.filter(user=user, status=Donation.Status.PENDING).exists()
    confirmed_donations_count = Donation.objects.filter(user=user, status=Donation.Status.APPROVED).count()
    latest_rejected_donation = (
        Donation.objects.filter(user=user, status=Donation.Status.REJECTED)
        .order_by("-created_at")
        .first()
    )
    donation_timeline = list(
        Donation.objects.filter(user=user)
        .select_related("method")
        .order_by("-created_at")[:5]
    )
    bot_username = (getattr(settings, "TELEGRAM_BOT_USERNAME", "") or "").strip().lstrip("@")
    bot_link = f"https://t.me/{bot_username}" if bot_username else ""
    bot_donate_link = f"https://t.me/{bot_username}?start=donat" if bot_username else ""
    household_members = []
    if user.household_id and user.is_supporter:
        household_members = list(
            user.household.members.order_by("date_joined").values_list("username", flat=True)
        )
    return render(
        request,
        "expenses/settings.html",
        {
            "user": user,
            "donation_methods": donation_methods,
            "pending_donation_exists": pending_donation_exists,
            "confirmed_donations_count": confirmed_donations_count,
            "latest_rejected_donation": latest_rejected_donation,
            "donation_timeline": donation_timeline,
            "bot_username": bot_username,
            "bot_link": bot_link,
            "bot_donate_link": bot_donate_link,
            "profile": profile,
            "household_members": household_members,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
@rate_limit_action("donation_moderation", max_requests=120, window=60)
def donation_moderation_view(request):
    """
    Staff uchun donat moderatsiya sahifasi (Web UI).
    """
    if not request.user.is_staff:
        messages.error(request, "Ushbu bo'lim faqat adminlar uchun.")
        return redirect("expenses:settings")

    from accounts.models import Donation

    if request.method == "POST":
        donation_id = request.POST.get("donation_id")
        donation_ids_raw = (request.POST.get("donation_ids") or "").strip()
        action = (request.POST.get("action") or "").strip().lower()
        reason = (request.POST.get("rejection_reason") or "").strip()
        ids = []
        if donation_ids_raw:
            for item in donation_ids_raw.split(","):
                item = item.strip()
                if item.isdigit():
                    ids.append(int(item))
        if donation_id and str(donation_id).isdigit():
            ids.append(int(donation_id))
        ids = sorted(set(ids))
        if not ids:
            messages.error(request, "Donat ID topilmadi.")
            return redirect("expenses:donation_moderation")
        donations = list(Donation.objects.select_related("user", "method").filter(pk__in=ids))
        if not donations:
            messages.error(request, "Donat topilmadi.")
            return redirect("expenses:donation_moderation")
        updated_count = 0
        if action == "approve":
            for donation in donations:
                donation.status = Donation.Status.APPROVED
                donation.rejection_reason = ""
                donation.save(update_fields=["status", "rejection_reason", "confirmed"])
                if donation.user and not donation.user.is_supporter:
                    donation.user.is_supporter = True
                    donation.user.save(update_fields=["is_supporter"])
                if donation.user and donation.user.telegram_id:
                    send_telegram_message(
                        donation.user.telegram_id,
                        "🎉 Donatingiz tasdiqlandi! Sizga Donater statusi berildi. Rahmat!",
                    )
                updated_count += 1
            messages.success(request, f"{updated_count} ta donat tasdiqlandi.")
        elif action == "reject":
            reject_reason = reason or "Chek bo'yicha ma'lumot aniqlashtirish talab qilindi."
            for donation in donations:
                donation.status = Donation.Status.REJECTED
                donation.rejection_reason = reject_reason
                donation.save(update_fields=["status", "rejection_reason", "confirmed"])
                if donation.user and donation.user.telegram_id:
                    send_telegram_message(
                        donation.user.telegram_id,
                        "❌ Donat tekshiruv natijasi: hozircha tasdiqlanmadi. Iltimos, chek screenshotini aniqroq ma'lumot bilan qayta yuboring.",
                    )
                updated_count += 1
            messages.info(request, f"{updated_count} ta donat rad etildi.")
        elif action == "pending":
            for donation in donations:
                donation.status = Donation.Status.PENDING
                donation.rejection_reason = ""
                donation.save(update_fields=["status", "rejection_reason", "confirmed"])
                updated_count += 1
            messages.success(request, f"{updated_count} ta donat qayta tekshiruvga o'tkazildi.")
        elif action == "ocr_extract":
            if len(donations) > 1:
                messages.error(request, "OCR faqat bitta donat uchun ishlaydi.")
                return redirect("expenses:donation_moderation")
            donation = donations[0]
            amount, err = _ocr_extract_amount_for_donation(donation)
            if err:
                messages.error(request, err)
                return redirect("expenses:donation_moderation")
            suggestion_key = OCR_SUGGESTION_CACHE_KEY.format(user_id=request.user.pk, donation_id=donation.pk)
            cache.set(suggestion_key, amount, timeout=3600)
            messages.success(request, f"OCR: summa {amount:,} so'm deb topildi. Tasdiqlashni bosing.")
        elif action == "ocr_apply":
            if len(donations) != 1:
                messages.error(request, "OCR apply uchun bitta donat tanlang.")
                return redirect("expenses:donation_moderation")
            donation = donations[0]
            suggestion_key = OCR_SUGGESTION_CACHE_KEY.format(user_id=request.user.pk, donation_id=donation.pk)
            amount = cache.get(suggestion_key) or 0
            if not amount:
                messages.error(request, "OCR taklif topilmadi yoki eskirgan.")
                return redirect("expenses:donation_moderation")
            donation.amount = amount
            note_tail = f"OCR amount={amount}"
            if note_tail not in (donation.note or ""):
                donation.note = f"{(donation.note or '').strip()} | {note_tail}".strip(" |")
            donation.save(update_fields=["amount", "note"])
            cache.delete(suggestion_key)
            messages.success(request, f"OCR summasi qo'llandi: {amount:,} so'm.")
        elif action == "ocr_discard":
            if len(donations) != 1:
                messages.error(request, "OCR discard uchun bitta donat tanlang.")
                return redirect("expenses:donation_moderation")
            donation = donations[0]
            suggestion_key = OCR_SUGGESTION_CACHE_KEY.format(user_id=request.user.pk, donation_id=donation.pk)
            cache.delete(suggestion_key)
            messages.info(request, "OCR taklifi bekor qilindi.")
        else:
            messages.error(request, "Noma'lum amal.")
        return redirect("expenses:donation_moderation")

    status_filter = (request.GET.get("status") or Donation.Status.PENDING).strip().lower()
    if status_filter not in {Donation.Status.PENDING, Donation.Status.APPROVED, Donation.Status.REJECTED, "all"}:
        status_filter = Donation.Status.PENDING

    q = (request.GET.get("q") or "").strip()
    qs = Donation.objects.select_related("user", "method").order_by("status", "-created_at")
    if q:
        from django.db.models import Q
        qs = qs.filter(
            Q(user__username__icontains=q)
            | Q(user__first_name__icontains=q)
            | Q(user__telegram_id__icontains=q)
            | Q(telegram_username_snapshot__icontains=q)
            | Q(note__icontains=q)
        )
    if status_filter != "all":
        qs = qs.filter(status=status_filter)

    status_counts = {
        "pending": Donation.objects.filter(status=Donation.Status.PENDING).count(),
        "approved": Donation.objects.filter(status=Donation.Status.APPROVED).count(),
        "rejected": Donation.objects.filter(status=Donation.Status.REJECTED).count(),
    }
    pending_sla_threshold = timezone.now() - timezone.timedelta(hours=24)
    pending_over_sla_count = Donation.objects.filter(
        status=Donation.Status.PENDING,
        created_at__lt=pending_sla_threshold,
    ).count()

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    ocr_suggestions = {}
    for d in page_obj:
        key = OCR_SUGGESTION_CACHE_KEY.format(user_id=request.user.pk, donation_id=d.pk)
        val = cache.get(key)
        if val:
            ocr_suggestions[d.pk] = val
    return render(
        request,
        "expenses/donation_moderation.html",
        {
            "page_obj": page_obj,
            "status_filter": status_filter,
            "q": q,
            "status_counts": status_counts,
            "pending_over_sla_count": pending_over_sla_count,
            "ocr_suggestions": ocr_suggestions,
            "ocr_suggestion_ids": list(ocr_suggestions.keys()),
            "status_choices": [
                ("pending", "Tekshiruvda"),
                ("approved", "Tasdiqlangan"),
                ("rejected", "Rad etilgan"),
                ("all", "Barchasi"),
            ],
        },
    )


@login_required
@require_http_methods(["POST"])
@rate_limit_action("export_supporter_report", max_requests=5, window=3600)
def export_supporter_report(request):
    """Donaterlar uchun Pro hisobotni Telegram botga yuboradi."""
    if not getattr(request.user, "is_supporter", False):
        messages.error(request, "Bu hisobot faqat Donater foydalanuvchilar uchun.")
        return redirect("expenses:settings")
    if not request.user.telegram_id:
        messages.error(request, "Telegram hisob topilmadi. Botda /start yuborib qayta urinib ko'ring.")
        return redirect("expenses:settings")
    today = timezone.now().date()
    cutoff = today - timezone.timedelta(days=365)
    from openpyxl import Workbook
    tmp_file_path = ""
    try:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="Donater Report")
        ws.append(["Sana", "Kategoriya", "Asl summa", "Valyuta", "UZS summa", "Izoh", "User"])
        qs = (
            Expense.objects.filter(user=request.user, date__gte=cutoff)
            .select_related("category", "user")
            .order_by("-date", "-created_at")
        )
        for e in qs.iterator(chunk_size=500):
            ws.append(
                [
                    str(e.date),
                    e.category.name if e.category else "",
                    float(e.original_amount),
                    e.currency,
                    int(e.amount),
                    e.note or "",
                    e.user.get_display_name(),
                ]
            )
        with NamedTemporaryFile(prefix=f"donater_report_{request.user.pk}_", suffix=".xlsx", delete=False) as tmp:
            tmp_file_path = tmp.name
        wb.save(tmp_file_path)
        caption = f"📊 Donater Pro hisobot ({today})"
        ok = send_telegram_document(request.user.telegram_id, tmp_file_path, caption=caption)
        if ok:
            messages.success(request, "Donater Pro hisobot Telegram botga yuborildi.")
        else:
            messages.error(request, "Hisobot botga yuborilmadi. Iltimos keyinroq qayta urinib ko'ring.")
    except Exception as e:
        logger.exception("export_supporter_report user_id=%s: %s", getattr(request.user, "pk", None), e)
        messages.error(request, "Donater Pro hisobotni tayyorlashda xatolik yuz berdi.")
    finally:
        if tmp_file_path and os.path.exists(tmp_file_path):
            try:
                os.remove(tmp_file_path)
            except OSError:
                pass
    return redirect("expenses:settings")


@login_required
@require_http_methods(["GET"])
def donation_moderation_photo(request, donation_id: int):
    """
    Staff uchun donation screenshot preview (Telegram file API orqali proxy).
    Bot token va file URL ni frontendga oshkor qilmaydi.
    """
    if not request.user.is_staff:
        return HttpResponse(status=403)

    from accounts.models import Donation

    donation = Donation.objects.filter(pk=donation_id).first()
    if not donation or not donation.screenshot_file_id:
        return HttpResponse("Rasm topilmadi", status=404)

    token = getattr(settings, "TELEGRAM_BOT_TOKEN", "")
    if not token:
        return HttpResponse("Bot token sozlanmagan", status=503)

    try:
        get_file_url = f"https://api.telegram.org/bot{token}/getFile"
        meta_resp = requests.get(get_file_url, params={"file_id": donation.screenshot_file_id}, timeout=10)
        meta_resp.raise_for_status()
        meta_data = meta_resp.json()
        file_path = ((meta_data or {}).get("result") or {}).get("file_path")
        if not file_path:
            return HttpResponse("Telegram fayl yo'li topilmadi", status=404)

        file_url = f"https://api.telegram.org/file/bot{token}/{file_path}"
        file_resp = requests.get(file_url, timeout=15)
        file_resp.raise_for_status()

        content_type = file_resp.headers.get("Content-Type", "image/jpeg")
        return HttpResponse(file_resp.content, content_type=content_type)
    except Exception as e:
        logger.warning("donation_moderation_photo error donation_id=%s: %s", donation_id, e)
        return HttpResponse("Rasmni yuklab bo'lmadi", status=502)
